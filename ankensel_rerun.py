import os
import glob
import openpyxl as px
import numpy as np
import pandas as pd
import datetime

dt_now = datetime.datetime.now()
print('starttime:',dt_now)

files=glob.glob(rf'C:\Users\m-nakagawa\Desktop\ankensheet\*.xlsm')
onlyfiles = next(os.walk(rf'C:\Users\m-nakagawa\Desktop\ankensheet'))[2]
endpointer = len(onlyfiles)

old_book = px.load_workbook(rf"C:\Users\m-nakagawa\Desktop\anken_old.xlsx",data_only=True)
sheet4 = old_book.active

old_last = old_book['Sheet'].max_row + 1

new_book = px.load_workbook(rf"C:\Users\m-nakagawa\Desktop\ankensel.xlsx",data_only=True)
#new_book = px.Workbook()
sheet3 = new_book.active
sheet3["A1"].value = "JOBNO"
sheet3["B1"].value = "売上金額(最新)"
sheet3["C1"].value = "営業利益(最新)"
sheet3["D1"].value = "利益率(最新)"
sheet3["E1"].value = "売上金額(計画)"
sheet3["F1"].value = "営業利益(計画)"
sheet3["G1"].value = "利益率(計画)"
sheet3["H1"].value = "案件名"
sheet3["I1"].value = "ＰＪ担当"
sheet3["J1"].value = "ランク"
sheet3["K1"].value = "チーム"

pointer = sheet3["L1"].value + 1
print("start:",pointer)
print("end:",endpointer)

for i in range(pointer,endpointer):
  file = files[i]
  workbook = px.load_workbook(file,data_only=True)
  sheet1 = workbook["案件シート"]
  job = sheet1["D9"]
  ankenmei = sheet1["D4"]
  tantou   = sheet1["S5"]
  ranku    = sheet1["M5"]
  kin1 = sheet1["G20"]
  rieki1 = sheet1["G36"]
  sheet2 = workbook["申請案件シート"]
  kin2 = sheet2["G20"]
  rieki2 = sheet2["G36"]
  print(i,"job:",job.value,"kin1:",kin1.value,"rieki1:",rieki1.value,"kin2:",kin2.value,"rieki2:",rieki2.value)
  ren = i + 2
  sheet3["L1"].value = i
  sheet3["A"+str(ren)].value = job.value
  sheet3["B"+str(ren)].value = kin1.value
  sheet3["C"+str(ren)].value = rieki1.value
  if kin2.value == 0 :
     sheet3["E"+str(ren)].value = kin1.value
     sheet3["F"+str(ren)].value = rieki1.value
  else:
     sheet3["E"+str(ren)].value = kin2.value
     sheet3["F"+str(ren)].value = rieki2.value
  sheet3["H"+str(ren)].value = ankenmei.value 
  sheet3["I"+str(ren)].value = tantou.value 
  sheet3["J"+str(ren)].value = ranku.value 
  for i2 in range(2,old_last):
    old_job      = sheet4["A"+str(i2)].value
    old_ankenmei = sheet4["H"+str(i2)].value
    old_teem     = sheet4["K"+str(i2)].value
    if   job.value == old_job :
       if ankenmei.value == old_ankenmei :
          sheet3["K"+str(ren)].value = old_teem  
  new_book.save(rf"C:\Users\m-nakagawa\Desktop\ankensel.xlsx")

dt_now = datetime.datetime.now()
print('endtime:',dt_now)
